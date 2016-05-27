import csv

import os.path
from optparse import make_option
from cells import CellBase, EmptyCell
from collections import OrderedDict

from django.core.management import BaseCommand, CommandError
from django.utils.six import with_metaclass
from django.utils import timezone
from django.db import transaction
from openpyxl import load_workbook


class ParserMetaclass(type):
    def __new__(cls, name, bases, attrs):
        attrs['fields'] = OrderedDict()
        built_fields = {}
        for field_name, obj in attrs.items():
            if isinstance(obj, CellBase):
                field = attrs[field_name]
                field.title = field_name
                built_fields[field_name] = field
        sorted_built_fields = OrderedDict(sorted(built_fields.items(), key=lambda x: x[1].creation_order))
        attrs['fields'].update(sorted_built_fields)
        return super(ParserMetaclass, cls).__new__(cls, name, bases, attrs)


class BaseParser(with_metaclass(ParserMetaclass, BaseCommand)):
    args = '<input_file>'
    option_list = BaseCommand.option_list+ (
        make_option(
            '--header',
            default=True,
            help='If there is header in the file'
        ),
        make_option(
            '-s',
            '--sheet',
            help='Set exact sheet for xlsx file'
        ),
        make_option(
            '--progress',
            default=False,
            help='Show progress bar? (progressbar package required)'
        ),
        make_option(
            '--failfast',
            default=False,
            help='Break parsing on first error-case?'
        ),
        make_option(
            '--dryrun',
            default=False,
            help='Do parsing without DB changes?'
        ),
        make_option(
            '--savestats',
            default=False,
            help='Do save statistics to file?'
        )
    )
    help = """
       Parse data
    """

    def __set_options(self, options):
        for key, value in options.items():
            if not key == 'sheet' and not isinstance(value, bool):
                parsed_value = True if value == 'True' else False
            else:
                parsed_value = value
            setattr(self, key, parsed_value)


    def __check_and_load_file(self, args):
        # We check, that file is exists and is valid and corresponding with options
        filename = args[0]
        self.filename = os.path.abspath(filename)
        name, extension = os.path.splitext(filename)
        if extension == '.csv':
            self.is_csv = True
        else:
            self.is_csv = False
        if not os.path.exists(self.filename):
            raise CommandError('Can\'t find given file: {}'.format(self.filename))
        if self.is_csv and hasattr(self, "sheet"):
            raise CommandError('Can\'t parse sheet for csv file!')
        elif not self.is_csv:
            try:
                self.work_book = load_workbook(filename, data_only=True)
            except BaseException:
                raise CommandError('Wrong file format. Supported are: .xlsx,.xlsm')

    def __check_and_load_sheet(self):
        # We will verify and load given sheet if it's exists or use first one.
        sheet_name = self.sheet
        if sheet_name is None:
            # WARNING Explicitly take first sheet.
            self.parsed_object = self.work_book.worksheets[0]
            return

        available_sheets = self.work_book.get_sheet_names()
        if not sheet_name in available_sheets:
            raise CommandError('Given sheet name `{}` not found. Existing: {} '.format(sheet_name, ", ".join(['`{}`'.format(x) for x in available_sheets])))
        self.parsed_object = self.work_book.get_sheet_by_name(sheet_name)

    def __initialize_progress_bar(self, total_rows):
        try:
            from progressbar import ProgressBar, Counter, Percentage, Bar
            widgets = ['Processed: ', Counter(), ' ', Percentage(), ' ', Bar()]
            return ProgressBar(widgets=widgets, maxval=total_rows).start()
        except ImportError:
            raise CommandError('No progressbar package found! Please, install it to track progress')

    def handle(self, *args, **options):
        self.parsed_successfully = 0
        self.parsed_unsuccessfully = 0
        self.skipped = 0
        self.__set_options(options)
        self.__check_and_load_file(args)
        if self.is_csv:
            self.parsed_object = open(self.filename, 'rb')
        else:
            self.__check_and_load_sheet()
        print 'Parsing file....'
        interim_data, total_rows = self.prepare_interim_data()
        if self.dryrun:
            transaction.set_autocommit(False)
        self.parse_data(interim_data, total_rows)
        self.after_parse_hook()
        self.parse_statistics()
        if self.dryrun:
            transaction.rollback()
            transaction.set_autocommit(True)

    def after_parse_hook(self):
        pass

    def prepare_interim_data(self):
        if self.is_csv:
            # Prepare progress bar data and generator for csv files
            object_generator = csv.reader(self.parsed_object, quotechar='"', delimiter=',')
            total_rows = sum(1 for line in object_generator)
            if self.header:
                self.parsed_object.seek(0)
                object_generator = csv.reader(self.parsed_object, quotechar='"', delimiter=',')
                total_rows -= 1
                next(object_generator)
        else:
            # Prepare progress bar data and generator for non-csv files
            row_offset = 1 if self.header else 0
            total_rows = self.parsed_object.max_row
            object_generator = self.parsed_object.iter_rows(row_offset=row_offset)
            total_rows -= 1
        return object_generator, total_rows

    def parse_data(self, object_generator, total_rows):
        if self.progress:
            pbar = self.__initialize_progress_bar(total_rows)
        for index, raw_row in enumerate(object_generator):
            if len(raw_row) != len(self.fields):
                raise CommandError('Incorrect parsed file! Stopping parsing! {} != {}'.format(len(raw_row),len(self.fields)))
            if self.is_csv:
                row = map(lambda s: s.strip(), raw_row)
            else:
                row = raw_row
            # TODO: invent great way to ignore last row when there is header
            if self.header and not self.is_csv and index == total_rows:
                continue
            self.process_row(row, index)
            if self.progress:
                pbar.update(index + 1)

    def process_row(self, row, row_number):
        if self.is_csv:
            row_raw_values = [cell for cell in row]
        else:
            row_raw_values = [cell.value for cell in row]
        if not any(row_raw_values):
            print "Blank line, SKIP"
            return None

        row_values = []
        row_errors = []
        row_values = dict()
        for index, cell in enumerate(row):

            option = self.fields.values()[index]
            if isinstance(option, EmptyCell):
                continue
            if self.is_csv:
                value = cell
            else:
                value = cell.value

            errors = option.validate(value)
            if errors:
                coordinates = index if self.is_csv else cell.coordinate
                if self.failfast:
                    raise CommandError('Errors in cell {}: {}'.format(coordinates, errors))
                else:
                    row_errors.append({coordinates: errors})
                continue
            value = option.normalize(value)

            try:
                handler = getattr(self, '{}_handler'.format(option.title))
                value = handler(value)
            except AttributeError:
                # In case we do not define handler at all.
                pass
            row_values[option.title] = value

        if row_errors:
            print "=" * 80
            print row_errors
            print "=" * 80
        try:
            row_handler = getattr(self, 'row')
        except AttributeError:
            # We do not perform any action with row itself.
            # We need to log, that we are skip row process for row, bla, bla.
            pass
        else:
            try:
                res = row_handler(row_values)
                if res is None:
                    res = self.success('Row {} parsed successfully'.format(row_number))
            except BaseException, e:
                res = self.failure('Error during parsing row {}: {}'.format(row_number, e))
            self.__process_result(res)

    def success(self, message):
        result_dict = {
            'status': 'Success',
            'message': message
        }
        return result_dict

    def failure(self, message):
        result_dict = {
            'status': 'Failure',
            'message': message
        }
        return result_dict

    def skip(self, message):
        result_dict = {
            'status': 'Skipped',
            'message': message
        }
        return result_dict

    def __process_result(self, res):
        status = res['status']
        if status == 'Success':
            self.parsed_successfully += 1
        elif status == 'Failure':
            self.parsed_unsuccessfully += 1
        elif status == 'Skipped':
            self.skipped += 1
        else:
            raise CommandError('Unexpected result returned')

    def parse_statistics(self, res):
        print 'Done!'
        print 'Successfully parsed {} items.'.format(self.parsed_successfully)
        print 'Failed to parse {} items.'.format(self.parsed_unsuccessfully)
        print 'Skipped {} items.'.format(self.skipped)
        if self.savestats:
            output_file_name = 'parse_statistics_' + timezone.now().strftime("%Y%m%d-%H%m") + '.txt'
            print 'Saving statistics to file: {}'.format(output_file_name)
            # TODO: implement real saving!
