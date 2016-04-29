import csv

import os.path
from optparse import make_option

from django.core.management import BaseCommand, CommandError
from progressbar import ProgressBar, Counter, Percentage, Bar
from openpyxl import load_workbook


class BaseParser(BaseCommand):
    args = '<input_file>'
    option_list = BaseCommand.option_list+ (
        make_option('--no-header',
            default=False,
            help='Update DB records related to existing rows.'),
        make_option('-s','--sheet',
            help='Update DB records related to existing rows.')
    )
    help = """
       Parse data
    """

    def __check_and_load_file(self, args, options):
        # We check, that file is exists and is valid and corresponding with options
        filename = args[0]
        self.filename = os.path.abspath(filename)
        name, extension = os.path.splitext(filename)
        if extension == '.csv':
            self.is_csv = True
        else:
            self.is_csv = False
        if not os.path.exists(self.filename):
            raise CommandError('Can\'t find given file: {}'.format(self.filename))
        if self.is_csv and options.get('sheet', None):
            raise CommandError('Can\'t parse sheet for csv file!')
        elif not self.is_csv:
            try:
                self.work_book = load_workbook(filename, data_only=True)
            except BaseException:
                raise CommandError('Wrong file format. Supported are: .xlsx,.xlsm')

    def __check_and_load_sheet(self, options):
        # We will verify and load given sheet if it's exists or use first one.
        sheet_name = options['sheet']
        if sheet_name is None:
            # WARNING Explicitly take first sheet.
            self.parsed_object = self.work_book.worksheets[0]
            return

        available_sheets = self.work_book.get_sheet_names()
        if not sheet_name in available_sheets:
            raise CommandError('Given sheet name `{}` not found. Existing: {} '.format(sheet_name, ", ".join(['`{}`'.format(x) for x in available_sheets])))
        self.parsed_object = self.work_book.get_sheet_by_name(sheet_name)

    def handle(self, *args, **options):
        self.__check_and_load_file(args, options)
        if self.is_csv:
            self.parsed_object = open(self.filename, 'rb')
        else:
            self.__check_and_load_sheet(options)
        print 'Parsing file....'
        self.parse_data(options)
        self.after_parse_hook()

    def after_parse_hook(self):
        pass

    def parse_data(self, options):
        widgets = ['Processed: ', Counter(), ' ', Percentage(), ' ', Bar()]
        if self.is_csv:
            # Prepare progress bar data and generator for csv files
            object_generator = csv.reader(self.parsed_object, quotechar='"', delimiter=',')
            total_rows = sum(1 for line in object_generator)
            if not options['no_header']:
                self.parsed_object.seek(0)
                object_generator = csv.reader(self.parsed_object, quotechar='"', delimiter=',')
                total_rows -= 1
                next(object_generator)
        else:
            # Prepare progress bar data and generator for non-csv files
            row_offset = 0 if options['no_header'] else 1
            total_rows = self.parsed_object.max_row
            object_generator = self.parsed_object.iter_rows(row_offset=row_offset)
        pbar = ProgressBar(widgets=widgets, maxval=total_rows).start()
        for index, raw_row in enumerate(object_generator):
            if self.is_csv:
                row = map(lambda s: s.strip(), raw_row)
            else:
                row = raw_row
            self.process_row(row)
            pbar.update(index + 1)

    def process_row(self, row):
        if self.is_csv:
            row_raw_values = [cell for cell in row]
        else:
            row_raw_values = [cell.value for cell in row]
        if not any(row_raw_values):
            print "Blank line, SKIP"
            return None

        row_values = []
        row_errors = []
        row_values = dict()
        for index, cell in enumerate(row):
            try:
                option = self.STRUCTURE[index]
                if self.is_csv:
                    value = cell
                else:
                    value = cell.value
            except IndexError:
                # Warning: skip cell processing as no related option.
                continue

            value = option.normalize(value)
            errors = option.validate(value)
            if errors:
                if self.is_csv:
                    row_errors.append({index: errors})
                else:
                    row_errors.append({cell.coordinate: errors})
                continue
            try:
                handler = getattr(self, '{}_handler'.format(option.title))
                value = handler(value)
            except AttributeError:
                # In case we do not define handler at all.
                pass
            row_values[option.title] = value

        if row_errors:
            print "=" * 80
            print row_errors
            print "=" * 80
        try:
            row_handler = getattr(self, 'row')
        except AttributeError:
            # We do not perform any action with row itself.
            # We need to log, that we are skip row process for row, bla, bla.
            pass
        else:
            row_handler(row_values)
